import openpyxl

# wb = openpyxl.load_workbook('14.TestSheet.xlsx')
# wb.sheetnames
# sheet = wb['Arkusz1']
# sheet
# type(sheet)
# sheet.title
# active_sheet = wb.active
# active_sheet

# a1 = sheet['A1']
# sheet['A1'].value

# 'Row %s, Column %s is %s' % (a1.row, a1.column, a1.value)

# sheet.cell(row=1, column=2) # get exact cell 

# sheet.max_row
# sheet.max_column

# ## Get rows from each colymns

# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')

# ## Get columns 

# list(sheet.columns)[1] # Get second column's cells.

# for cellObj in list(sheet.columns)[1]:
#         print(cellObj.value)

# ## Creating new sheet

# wb.create_sheet(index=0, title='First Sheet')
# wb.sheetnames
# del wb['First Sheet']

# Write value in cell

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!' # Edit the cell's value.
sheet['A1'].value

# Change text format

from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font.
sheet['A1'].font = italic24Font # Apply the font to A1.
sheet['A1'] = 'Hello, world!'
wb.save('styles.xlsx')

# Change formulas

sheet['B9'] = '=SUM(B1:B8)'

# Change H an W of rows/collumns

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
# Set the height and width:
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# Merging cells

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5') # Merge these two cells.
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

# Creating charts

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column A
     sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1,
max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')